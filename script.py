import re
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows

def try_again(ls, ind):
	try:
		return ls[ind]
	except:
		return None


def getDescription(a1, a2, a3, a4):
	a2 = [f'<li>{x}</li>' for x in a2]
	a2 = ''.join(a2)
	a2 = f'<ul>{a2}</ul>'.replace('’', "'")
	a4 = [f'<li>{x}</li>' for x in a4]
	a4 = ''.join(a4)
	a4 = f'<div><span>Features:</span> <ul>{a4}</ul></div>'.replace('’', "'")
	a1 = f'<div>{a1}</div>'
	a3 = f'<div><span>Style #:</span><span>{a3}</span></div>'
	desc = f'{a1}{a2}{a4}{a3}'
	return desc


def get_variables(dt:dict,vendor):
	stock=dt["stock"]
	title_gender = dt["gender"]["title_gender"]
	title = dt["title"]
	title = title.title().replace("New Balance Numeric", "").replace(
		"Nb Numeric ", "").replace("New Balance ", "").replace("Kids", "")
	title = remove_double_spaces(title)
	color = dt["color"].title()
	main_color = color.replace("/", " ").strip()
	# TODO k-siwss only
	for s in ["-M","-m","-W","-w"]:
		if main_color.find(s)!=-1:
			main_color=main_color.replace(s,"")
			break
	color=main_color.split(" ")[0]
	type_ = dt["type"]

	if 'new_title' not in dt.keys():
		new_title = f'{vendor} {title_gender} {title} {main_color} {type_}'.replace(
			'with ', '').replace('| ', '').title().replace("Men'S", "Men's").replace("Women'S", "Women's")
		new_title = remove_double_spaces(new_title)
	else:
		new_title=dt['new_title']
	
	if 'handle' not in dt.keys():
		handle=new_title.lower().replace(' ', '-')
		handle=handle.replace("'","")
	else:
		handle=dt['handle']

	return dt["url"], title_gender, dt["gender"]["gender"], dt["gender"]["age_group"], title, dt["style_code"], color, dt["sizes"], dt["price"], dt["cost"], dt["description"], dt["features"], dt["bullet_points"], dt["widths"], dt["images"], dt["category"], dt["weight"], new_title,type_,main_color,stock,handle


def remove_double_spaces(text):
	return re.sub(r' +', ' ', text)


def get_workbook(sheet_name):
	file_path = 'Templates.xlsx' # Replace with the path to your existing Excel file
	workbook = openpyxl.load_workbook(file_path)
	sheet = workbook[sheet_name]
	new_workbook = openpyxl.Workbook()
	new_sheet = new_workbook.active
	# for row in sheet.iter_rows(values_only=True):
	#     new_sheet.append(row)
	for row in sheet.iter_rows():
		new_row = []
		for cell in row:
			new_cell = new_sheet.cell(
				row=cell.row, column=cell.column, value=cell.value)
			new_cell.font = copy(cell.font)
			new_cell.border = copy(cell.border)
			new_cell.fill = copy(cell.fill)
			new_cell.number_format = copy(cell.number_format)
			new_cell.alignment = copy(cell.alignment)
			new_cell.protection = copy(cell.protection)
		# Close both workbooks when done
	workbook.close()
	return new_workbook


# Shopify
def get_shopify_product_data(ls,vendor,workbook):
	sheet=workbook["Shopify"]
	qty_sheet=workbook["Shopify.Qty"]
	products_data = []
	for dt in ls:
		# Extracting data from dict
		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle= get_variables(dt,vendor)
		color=main_color
		img_counter = 0
		tags = [vendor]
		tags.append(title_gender)
		tags.append(title)
		# tags.extend(color.replace('with ', '').split(' '))
		tags.append(type_)
		tags.append(style_code)
		tags.extend(sizes)
		tags.append(price)
		tags.append(cost)
		tags.append(color)
		tags = list(map(lambda x: str(x), tags))
		tags = ','.join(tags)
		data = {}
		data['Url'] = url
		data['Handle'] = handle
		data['Title'] = new_title
		data['Body (HTML)'] = getDescription(
			description, bullet_point, style_code, features).replace('’', "'")
		data['Vendor'] = vendor
		data['Standardized Product Type'] = category
		data['Custom Product Type'] = f'{title_gender} {type_}'
		data['Tags'] = tags.replace('| ', '')
		data['Published'] = 'TRUE'
		data['Option1 Name'] = 'Select a Size:'
		data['Option1 Value'] = sizes[0]
		data['Option2 Name'] = 'Color:'
		data['Option2 Value'] = str(color).split()[0]
		if len(widths) > 0:
			data['Option3 Name'] = 'Width:'
			data['Option3 Value'] = widths[0].strip()
		else:
			data['Option3 Name'] = ''
			data['Option3 Value'] = ''
		data['Variant SKU'] = stock[0]["SKU"]
		data['Variant Grams'] = weight
		data['Variant Inventory Tracker'] = 'shopify'
		data['Variant Inventory Qty'] = stock[0]["Quantity"]
		data['Variant Inventory Policy'] = 'deny'
		data['Variant Fulfillment Service'] = 'manual'
		data['Variant Price'] = price
		data['Variant Compare At Price'] = ''
		data['Variant Requires Shipping'] = 'TRUE'
		data['Variant Taxable'] = 'TRUE'
		data['Variant Barcode'] = stock[0]["Upc"]
		data['Image Src'] = ''
		data['Image Position'] = ''
		data['Image Alt Text'] = ''
		data['Gift Card'] = 'FALSE'
		data['SEO Title'] = new_title
		seo_description=BeautifulSoup(description, "html.parser").get_text(strip=False).strip("\n")
		data['SEO Description'] = f"{seo_description}\n"
		if len(bullet_point)>0:
			seo_description+=",".join(bullet_point)
		if len(features)>0:
			seo_description+=",".join(features)
		seo_description+=f"Style # {style_code}"
		# data['SEO Description'] = f'{description}{",".join(bullet_point)}{",".join(features)}\nStyle # {style_code}'
		data['Google Shopping / Google Product Category'] = category
		data['Google Shopping / Gender'] = gender
		data['Google Shopping / Age Group'] = age_group
		data['Google Shopping / MPN'] = stock[0]["Upc"]
		data['Google Shopping / AdWords Grouping'] = ''
		data['Google Shopping / AdWords Labels'] = ''
		data['Google Shopping / Condition'] = 'New'
		data['Google Shopping / Custom Product'] = ''
		data['Google Shopping / Custom Label 0'] = data['Option2 Value']
		# data['Google Shopping / Custom Label 0'] = color
		data['Google Shopping / Custom Label 1'] = ''
		data['Google Shopping / Custom Label 2'] = ''
		data['Google Shopping / Custom Label 3'] = ''
		data['Google Shopping / Custom Label 4'] = ''
		data['Variant Image'] = ''
		data['Variant Weight Unit'] = 'g'
		data['Variant Tax Code'] = ''
		data['Cost per item'] = cost
		data['Status'] = 'Active'
		products_data.append(data)
		qty_sheet.append([data["Handle"],data["Title"],data["Option1 Name"],data["Option1 Value"],data["Option2 Name"],data["Option2 Value"],data["Option3 Name"],data["Option3 Value"],data['Variant SKU'],"","","FSS Warehouse",stock[0]["Quantity"]])
		for st in stock[1:]:
			img = try_again(images, img_counter)
			data = {}
			data['Url'] = url
			data['Handle'] = handle
			data['Title'] = ''
			data['Body (HTML)'] = ''
			data['Vendor'] = ''
			data['Standardized Product Type'] = ''
			data['Custom Product Type'] = ''
			data['Tags'] = ''
			data['Published'] = ''
			data['Option1 Name'] = ''
			data['Option1 Value'] = st["size"]
			data['Option2 Name'] = ''
			data['Option2 Value'] = str(color).split()[0]
			if len(widths) > 0:
				data['Option3 Name'] = ''
				data['Option3 Value'] = widths[0].strip()
			else:
				data['Option3 Name'] = ''
				data['Option3 Value'] = ''
			data['Variant SKU'] = st["SKU"]
			data['Variant Grams'] = weight
			data['Variant Inventory Tracker'] = 'shopify'
			data['Variant Inventory Qty'] = st["Quantity"]
			data['Variant Inventory Policy'] = 'deny'
			data['Variant Fulfillment Service'] = 'manual'
			data['Variant Price'] = price
			data['Variant Compare At Price'] = ''
			data['Variant Requires Shipping'] = 'TRUE'
			data['Variant Taxable'] = 'TRUE'
			data['Variant Barcode'] = st["Upc"]
			if img:
				data['Image Src'] = img
				data['Image Position'] = img_counter + 1
				data['Image Alt Text'] = new_title
				img_counter += 1
			else:
				data['Image Src'] = ''
				data['Image Position'] = ''
				data['Image Alt Text'] = ''

			data['Gift Card'] = ''
			data['SEO Title'] = ''
			data['Google Shopping / Google Product Category'] = category
			data['Google Shopping / Gender'] = gender
			data['Google Shopping / Age Group'] = age_group
			data['Google Shopping / MPN'] = st["Upc"]
			data['Google Shopping / AdWords Grouping'] = ''
			data['Google Shopping / AdWords Labels'] = ''
			data['Google Shopping / Condition'] = 'New'
			data['Google Shopping / Custom Product'] = ''
			# data['Google Shopping / Custom Label 0'] = color
			data['Google Shopping / Custom Label 0'] = data['Option2 Value']
			data['Google Shopping / Custom Label 1'] = ''
			data['Google Shopping / Custom Label 2'] = ''
			data['Google Shopping / Custom Label 3'] = ''
			data['Google Shopping / Custom Label 4'] = ''
			data['Variant Image'] = ''
			data['Variant Weight Unit'] = 'g'
			data['Variant Tax Code'] = ''
			data['Cost per item'] = cost
			data['Status'] = 'Active'
			qty_sheet.append([data["Handle"],data["Title"],data["Option1 Name"],data["Option1 Value"],data["Option2 Name"],data["Option2 Value"],data["Option3 Name"],data["Option3 Value"],data['Variant SKU'],"","","FSS Warehouse",st["Quantity"]])
			products_data.append(data)
	# for p in products_data:
	# 	sheet.append(p)
	# for row in sheet.iter_rows(): # Start from the second row (skip headers)
	# 	for cell in row:
	# 		cell.value = None
	sheet.delete_rows(0, sheet.max_row) # Start from the second row to the last row
	# Delete all columns
	sheet.delete_cols(1, sheet.max_column)
	df=pd.DataFrame(products_data).drop('Url', axis=1)
	for row in dataframe_to_rows(df, index=False):
		sheet.append(row)
	return df
	# pd.DataFrame(products_data).drop_duplicates(subset=['Variant SKU']).drop('Url', axis=1).to_csv(f"{vendor} Shopify Product.csv", index=False)

# Amazon related
def get_amazon_product_row(dt, variation=True):
	data_dict = {
		"Product Type": "",
		"Seller SKU": "",
		"Brand": "",
		"Update Delete": "Update",
		"Product Name": "",
		"Product ID": "",
		"Product ID Type": "",
		"Product Description": "",
		"Item Type1": "",
		"Style Number": "",
		"Sole Material": "",
		"Heel Type": "",
		"Model Year": "",
		"Heel Height": "no-heel",
		"Closure Type": "lace-up",
		"Model Name": "",
		"Outer Material Type": "suede",
		"Manufacturer Part Number": "",
		"Manufacturer": "",
		"Heel Height Unit Of Measure": "",
		"Unit of measure of the height of the heel": "",
		"Footwear Size System": "US_Footwear_Size_System",
		"Shoe Size Age Group": "",
		"Shoe Size Gender": "",
		"Shoe Size Class": "Numeric",
		"Shoe Size Width": "Medium",
		"Shoe Size": "",
		"Shoe Size To Range (If Range)": "",
		"Shoe Size Gender (If Unisex)": "",
		"Shoe Size Width (If Unisex)": "",
		"Shoe Size (If Unisex)": "",
		"Shoe Size To Range (If Unisex and Range)": "",
		"Your Price": "",
		"Quantity": "",
		"Target Gender": "",
		"Age Range Description": "",
		"Other Image URL1": "",
		"Other Image URL2": "",
		"Other Image URL3": "",
		"Other Image URL4": "",
		"Other Image URL5": "",
		"Other Image URL6": "",
		"Other Image URL7": "",
		"Other Image URL8": "",
		"Main Image URL": "",
		"Swatch Image URL": "",
		"Parentage": "",
		"Parent SKU": "",
		"Relationship Type": "",
		"Variation Theme": "Size",
		"Package Level": "",
		"package_contains_quantity": "",
		"package_contains_identifier": "",
		"Style": "",
		"Bullet Point1": "",
		"Bullet Point2": "",
		"Bullet Point3": "",
		"Bullet Point4": "",
		"Bullet Point5": "",
		"Search Terms": "",
		"Style Keyword": "",
		"Toe Style": "close-toe",
		"Arch Type": "",
		"Team Name": "",
		"Boot Opening Circumference": "",
		"Color": "",
		"Color Map": "",
		"Size": "",
		"is-stain-resistant": "",
		"Material Fabric": "",
		"Country/Region as Labeled": "",
		"Fur Description": "",
		"Pattern Style": "",
		"Water Resistance Level": "",
		"Occasion Lifestyle": "",
		"Special Features1": "",
		"Special Features2": "",
		"Special Features3": "",
		"Special Features4": "",
		"Special Features5": "",
		"Platinum Keywords1": "",
		"Platinum Keywords2": "",
		"Platinum Keywords3": "",
		"Platinum Keywords4": "",
		"Platinum Keywords5": "",
		"Season of the Product": "",
		"Collection": "",
		"Target Audience1": "",
		"Target Audience2": "",
		"Target Audience3": "",
		"Target Audience4": "",
		"Target Audience5": "",
		"Catalog Number": "",
		"Minimum Circumference Unit of Measure": "",
		"Item Type2": "fashion-sneakers",
		"Athlete": "",
		"Pronation Correction": "",
		"Special Size Type": "",
		"shaft_style_type": "",
		"Product Lifecycle Supply Type1": "",
		"Product Lifecycle Supply Type2": "",
		"Product Lifecycle Supply Type3": "",
		"Product Lifecycle Supply Type4": "",
		"Product Lifecycle Supply Type5": "",
		"Item Booking Date": "",
		"Cleat Description": "",
		"Cleat Material Type": "",
		"Strap Type": "",
		"Duration Unit": "",
		"Certifying Authority Name": "",
		"character": "",
		"Supplier Package Type": "Retail Box",
		"Duration": "",
		"flash_point_unit_of_measure": "",
		"Shaft Diameter Unit": "",
		"Insole Type": "",
		"Built-in Light": "",
		"Shaft Diameter": "",
		"Shoe Safety Code ISO-20345": "",
		"Geographic Jurisdiction": "",
		"Wine Label (Fanciful Name)": "",
		"Surface Recommendation": "",
		"Shipping Weight": "",
		"Website Shipping Weight Unit Of Measure": "",
		"Shoe Height Map": "",
		"Shaft Height": "",
		"Platform Height": "",
		"Item Length Unit Of Measure": "",
		"Item Length1": "",
		"Item Width": "",
		"Item Height": "",
		"Width": "",
		"Item Display Height Unit Of Measure": "",
		"Display Height": "",
		"Maximum Girth Size": "",
		"Shoe Width": "",
		"Item Height Unit Of Measure": "",
		"shoe_width_unit_of_measure": "",
		"Shaft Height Unit Of Measure": "",
		"Item Width Unit Of Measure": "",
		"Item Display Width Unit Of Measure": "",
		"Maximum Circumference Unit of Measure": "",
		"Item Length2": "",
		"Item Display Length Unit Of Measure": "",
		"Fulfillment Center ID": "",
		"Package height": "",
		"Package Width": "",
		"Package Length": "",
		"Package Length Unit Of Measure": "",
		"Package Weight": "",
		"Package Weight Unit Of Measure": "",
		"Package Height Unit Of Measure": "",
		"Package Width Unit Of Measure": "",
		"Compliance Regulation Type1": "",
		"Compliance Regulation Type2": "",
		"Compliance Regulation Type3": "",
		"Compliance Regulation Type4": "",
		"Compliance Regulation Type5": "",
		"Regulatory Identification1": "",
		"Regulatory Identification2": "",
		"Regulatory Identification3": "",
		"Regulatory Identification4": "",
		"Regulatory Identification5": "",
		"Fabric Type": "",
		"Import Designation": "",
		"CPSIA Warning Description": "",
		"Cpsia Warning": "",
		"item_weight_unit_of_measure": "",
		"Item Weight": "",
		"Is this product a battery or does it utilize batteries?": "",
		"Batteries are Included": "",
		"Battery composition": "",
		"Battery type/size1": "",
		"Battery type/size2": "",
		"Battery type/size3": "",
		"Number of batteries": "",
		"Battery weight (grams)": "",
		"battery_weight_unit_of_measure": "",
		"Number of Lithium Metal Cells": "",
		"Number of Lithium-ion Cells": "",
		"Lithium Battery Packaging": "",
		"Watt hours per battery": "",
		"lithium_battery_energy_content_unit_of_measure": "",
		"Lithium content (grams)": "",
		"Applicable Dangerous Goods Regulations1": "",
		"Applicable Dangerous Goods Regulations2": "",
		"Applicable Dangerous Goods Regulations3": "",
		"Applicable Dangerous Goods Regulations4": "",
		"Applicable Dangerous Goods Regulations5": "",
		"UN number": "",
		"Safety Data Sheet (SDS) URL": "",
		"Volume": "",
		"item_volume_unit_of_measure": "",
		"Country/Region of Origin": "",
		"Compliance Certification Status": "",
		"Flash point (°C)?": "",
		"legal_compliance_certification_date_of_issue": "",
		"Safety Warning": "",
		"Compliance Certification Value": "",
		"legal_compliance_certification_expiration_date": "",
		"Regulatory Organization Name": "",
		"Legal Disclaimer Description": "",
		"Material/Fabric Regulations1": "",
		"Material/Fabric Regulations2": "",
		"Material/Fabric Regulations3": "",
		"Legal Compliance Certification":"",
		"Categorization/GHS pictograms (select all that apply)1": "",
		"Categorization/GHS pictograms (select all that apply)2": "",
		"Categorization/GHS pictograms (select all that apply)3": "",
		"California Proposition 65 Warning Type": "",
		"California Proposition 65 Chemical Names": "",
		"Additional Chemical Name1": "",
		"Additional Chemical Name2": "",
		"Additional Chemical Name3": "",
		"Additional Chemical Name4": "",
		"Pesticide Marking1": "",
		"Pesticide Marking2": "",
		"Pesticide Marking3": "",
		"Pesticide Registration Status1": "",
		"Pesticide Registration Status2": "",
		"Pesticide Registration Status3": "",
		"Pesticide Certification Number1": "",
		"Pesticide Certification Number2": "",
		"Pesticide Certification Number3": "",
		"Radio Frequency Emission & Authorization Status": "",
		"SDoC Contact Email Address": "",
		"SDOC Contact US Phone Number": "",
		"SDoC Contact Name": "",
		"temp1":"", # FIXME hardcoded actual key look in template
		"temp2":"",
		"temp3":"",
		"FCC ID": "",
		"SDoC Contact US Mailing Address": "",
		"List Price": "",
		"Item Condition": "New",
		"Condition Note": "",
		"Currency": "USD",
		"Product Tax Code": "",
		"Handling Time": "3",
		"Launch Date": "",
		"Release Date": "",
		"Restock Date": "",
		"Sale Price": "",
		"Sale Start Date": "",
		"Sale End Date": "",
		"Max Aggregate Ship Quantity": "",
		"Item Package Quantity": "",
		"Offering Can Be Gift Messaged": "",
		"Is Gift Wrap Available": "",
		"Is Discontinued by Manufacturer?": "",
		"Offer End Date": "",
		"Shipping-Template": "",
		"Minimum Advertised Price": "",
		"Number of Items": "",
		"Max Order Quantity": "",
		"Offer Start Date": "",
		"Business Price": "",
		"Quantity Price Type": "",
		"Quantity Lower Bound 1": "",
		"Quantity Price 1": "",
		"Quantity Lower Bound 2": "",
		"Quantity Price 2": "",
		"Quantity Lower Bound 3": "",
		"Quantity Price 3": "",
		"Quantity Lower Bound 4": "",
		"Quantity Price 4": "",
		"Quantity Lower Bound 5": "",
		"Quantity Price 5": "",
		"Progressive Discount Type": "",
		"Progressive Discount Lower Bound 1": "",
		"Progressive Discount Value 1": "",
		"Progressive Discount Lower Bound 2": "",
		"Progressive Discount Value 2": "",
		"Progressive Discount Lower Bound 3": "",
		"Progressive Discount Value 3": "",
		"National Stock Number": "",
		"United Nations Standard Products and Services Code": "",
		"Pricing Action": ""
	}
	data_dict.update(dt)
	return list(data_dict.values())


def get_amazon_quantity_row(dt, variation=True):
	data_dict = {
		"sku": "",
		"price": "",
		"minimum-seller-allowed-price": "",
		"maximum-seller-allowed-price": "",
		"quantity": "",
		"leadtime-to-ship": "",
		"fulfillment-channel": ""
	}
	if variation:
		for key in data_dict:
			if key != "Max dispatch time":
				data_dict[key] = ""
	data_dict.update(dt)
	return list(data_dict.values())

# Walmart related


def get_walmart_row(dt, variation=True):
	data_dict = {
		"Style Code": "", # Added
		"Product ID Type": "UPC",
		"Product ID": "",
		"Product Name": "",
		"Brand": "",
		"Selling Price": "",
		"Shipping Weight (lbs)": "",
		"Site Description": "",
		"Main Image URL": "",
		"Additional Image URL (+)": "",
		"Additional Image URL 1 (+)": "",
		"Additional Image URL 2 (+)": "",
		"Additional Image URL 3 (+)": "",
		"California Prop 65 Warning Text": "",
		"Small Parts Warning Code (+)": "",
		"Country of Origin - Textiles": "",
		"Gender": "",
		"Shoe Size": "",
		"Age Group (+)": "",
		"Shoe Category": "",
		"Color Category (+)": "",
		"Key Features (+)": "",
		"Key Features 1 (+)": "",
		"Key Features 2 (+)": "",
		"MSRP": "",
		"Manufacturer Name": "",
		"Count Per Pack": "",
		"Total Count": "",
		"Shoe Width": "",
		"Shoe Size System": "US Men's",
		"Color (+)": "",
		"Pattern (+)": "",
		"Material (+)": "",
		"Activity (+)": "",
		"Measure1":"",
		"Unit1": "",
		"Shoe Style": "",
		"Shoe Attire": "",
		"Shoe Closure": "",
		"Collection": "",
		"Secondary Material (+)": "",
		"Shoe Sole Material": "",
		"Insole Feature (+)": "",
		"Fabric Material Name": "",
		"Fabric Material Percentage": "",
		"Fabric Care Instructions (+)": "",
		"Measure2":"",
		"Unit2": "",
		"Footwear Heel Style": "",
		"Toe Style": "",
		"Embellishment Type (+)": "",
		"Measure3": "",
		"Unit3": "",
		"Measure4": "",
		"Unit4": "",
		"Boot Shaft Style (+)": "",
		"Boot Shaft Width Style (+)": "",
		"Season (+)": "",
		"Occasion (+)": "",
		"Recommended Locations (+)": "",
		"Is Water-Resistant": "",
		"Is Orthopedic": "",
		"Health Concerns (+)": "",
		"Size": "",
		"Character (+)": "",
		"Sports Team (+)": "",
		"Sports League (+)": "",
		"Athlete (+)": "",
		"Autographed by": "",
		"Brand License (+)": "",
		"Warranty Text": "",
		"Warranty URL": "",
		"Additional Features (+)": "",
		"Additional Product Attribute Name": "",
		"Additional Product Attribute Value": "",
		"Season Year": "",
		"Season Code": "",
		"Variant Group ID": "",
		"Variant Attribute Names (+)": "shoeSize",
		"Is Primary Variant": "",
		"Swatch Variant Attribute": "shoeSize",
		"Swatch Image URL": "",
		"Restriction Type": "",
		"States": "",
		"ZIP Codes": "",
		"Fulfillment Lag Time": "",
		"PPU Quantity of Units": "",
		"PPU Unit of Measure": "",
		"Contains Electronic Component?": "",
		"Contained Battery Type": "",
		"Contains Chemical, Aerosol or Pesticide?": "",
		"Multipack Quantity": "",
		"Ships in Original Packaging": "",
		"Must ship alone?": "",
		"Local Delivery Depth (in)": "",
		"Local Delivery Width (in)": "",
		"Local Delivery Height (in)": "",
		"Local Delivery Weight (lbs)": "",
		"Site Start Date": "",
		"Site End Date": "",
		"Additional Offer Attribute Name": "",
		"Additional Offer Attribute Value": "",
		"External Product ID Type": "",
		"External Product ID": "",
		"Product Id Update": "",
		"SKU Update": ""
	}
	# if variation:
	#     for key in data_dict:
	#         if key!="Max dispatch time":
	#             data_dict[key] = ""
	data_dict.update(dt)
	return ["", "", ""]+list(data_dict.values())

color_dict={}
for row in pd.read_excel("ColorLookupWalmart.xlsx").values.tolist():
	color_dict[row[0]]=row[1]

def get_walmart_quantity_row(dt, variation=True):
	data_dict = {
		"SKU*": "",
		"New Quantity*": "",
		"Fulfillment Center ID": "" 
	}
	# if variation:
	#     for key in data_dict:
	#         if key!="Max dispatch time":
	#             data_dict[key] = ""
	data_dict.update(dt)
	# return ["","",""]+list(data_dict.values())
	return list(data_dict.values()) 


def get_walmart_qunatity_data(ls,sheet):
	# file_path = 'Templates/WalmartQuantityTemplate.xlsx'
	# workbook = openpyxl.load_workbook(file_path)
	# sheet = workbook.active
	for dt in ls:
		for s in dt["stock"]:
			m_dt = {
				"SKU*" :s["SKU"],
				"New Quantity*" :s["Quantity"],
				"Fulfillment Center ID": "10001446565",
			}
			ls = get_walmart_quantity_row(m_dt)
			sheet.append(ls)
		# break
		# workbook.save('Walmart Quantity Data.xlsx')
		# workbook.close()

def get_ebay_quantity_row(dt, variation=True):
	data_dict = {
		"Action": "Revise",
		"Item number": "155422453475",
		"Title": "",
		"Listing site": "",
		"Currency": "",
		"Start price": "",
		"Buy It Now price": "",
		"Available quantity": "",
		"Relationship": "",
		"Relationship details": "",
		"Custom label (SKU)": "" 
	}
	if variation:
		for key in data_dict:
			if key!="Max dispatch time":
				data_dict[key] = ""
	data_dict.update(dt)
	return list(data_dict.values())


def get_ebay_quantity_data(ls,vendor,sheet):
	for dt in ls:
		# Extracting data from dict
		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle = get_variables(dt,vendor)
		m_dt = {
			"Title": new_title,
			"Listing site": "US",
			"Currency":"USD",
			"Start price": price,
			"Start price": '',
			"Relationship": "",
			"Relationship details": f"US Shoe Size={';'.join(map(str, sizes))}",
		}
		ls = get_ebay_quantity_row(m_dt, False)
		sheet.append(ls)
		for s in dt["stock"]:
			m_dt = {
				# "Title": new_title,
				"Start price": price,
				"Available quantity": s["Quantity"],
				"Relationship": "Variation",
				"Relationship details": f"US Shoe Size={s['size']}",
				"Custom label (SKU)": s["SKU"],
			}
			ls = get_ebay_quantity_row(m_dt)
			sheet.append(ls)


def get_ebay_row(dt, variation=True):
	# postal_code="85373"
	postal_code="90017"
	data_dict = {
		"*Action(SiteID=US|Country=US|Currency=USD|Version=1193)": "Add",
		"Custom label (SKU)": "",
		"Category ID": "15709",
		"Category name": "",
		"Title": "",
		"Relationship": "",
		"Relationship details": "",
		"P:UPC": "",
		"P:ISBN": "",
		"P:EAN": "",
		"P:EPID": "",
		"Start price": "",
		"Quantity": "",
		"Item photo URL": "",
		"Condition ID": "1000-New with box",
		"Description": "",
		"Format": "FixedPrice",
		"Duration": "GTC",
		"Buy It Now price": "",
		"Paypal accepted": "",
		"Paypal email address": "",
		"Immediate pay required": "",
		"Payment instructions": "",
		"PostalCode": postal_code,
		"Shipping service 1 option": "",
		"Shipping service 1 cost": "",
		"Shipping service 1 priority": "",
		"Shipping service 2 option": "",
		"Shipping service 2 cost": "",
		"Shipping service 2 priority": "",
		"Max dispatch time": "3",
		"WeightUnit": "lb",
		"WeightMajor": "3",
		"WeightMinor": "",
		"Returns accepted option": "",
		"Returns within option": "",
		"Refund option": "",
		"Return shipping cost paid by": "",
		"Shipping profile name": "*USPS Priority free, 3 business days",
		"Return profile name": "*Returns Accepted, International No Returns",
		"Payment profile name": "*ManagedPayments",
		"TakeBackPolicyID": "",
		"Regional TakeBackPolicies": "",
		"ProductCompliancePolicyID": "",
		"Regional ProductCompliancePolicies": "",
		"Hazmat Pictograms": "",
		"Hazmat SignalWord": "",
		"Hazmat Statements": "",
		"Hazmat Component": "",
		"EnergyEfficiencyLabel ImageURL": "",
		"EnergyEfficiencyLabel ImageDescription": "",
		"EnergyEfficiencyLabel ProductInformationSheet": "",
		"EcoParticipationFee": "",
		"RepairScore": "",
		"C:Brand": "New Balance Numeric",
		"C:Type": "Sneaker",
		"C:Size": "",
		"C:Style": "Athletic",
		"C:Outer Shell Material": "",
		"C:Size Type": "",
		"C:Department": "",
		"C:Color": "",
		"C:Upper Material": "Suede/Canvas",
		"C:Sole Material": "Rubber",
		"C:Shoe Width": "Medium",
		"C:Inseam": "",
		"C:Sleeve Length": "",
		"C:Lining Material": "",
		"C:Insulation Material": "",
		"C:Theme": "",
		"C:Fabric Type": "",
		"C:Features": "",
		"C:Pattern": "",
		"C:Accents": "",
		"C:Performance/Activity": "",
		"C:Model": "",
		"C:Product Line": "",
		"C:Country/Region of Manufacture": "",
		"C:Closure": "Lace Up",
		"C:Occasion": "Casual",
		"C:Vintage": "",
		"C:Waist Size": "",
		"C:Material": "",
		"C:Fabric Wash": "",
		"C:Leg Style": "",
		"C:Front Type": "",
		"C:Character": "",
		"C:Neckline": "",
		"C:Character Family": ""
	}
	if variation:
		for key in data_dict:
			# TODO:
			if key != "Max dispatch time" and key != "WeightUnit" and key !="WeightMajor":
				data_dict[key] = ""
	data_dict.update(dt)
	return list(data_dict.values())

def get_walmart_product_data(ls,vendor,workbook):
	qty_sheet=workbook["Walmart.Qty"]
	get_walmart_qunatity_data(ls,qty_sheet)
	sheet=workbook["Walmart"]
	# Replace with the path to your existing Excel file
	# file_path = 'TemplateEbayProduct.xlsx'
	# Replace with the path to your existing Excel file
	# if not sheet:
	# 	file_path = 'Templates/WalmartProductTemplate.xlsx'
	# 	workbook = openpyxl.load_workbook(file_path)
	# 	sheet = workbook.active
	for dt in ls:
		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle = get_variables(dt,vendor)



		try:
			color=color_dict[color]
		except:
			pass
		age_group=age_group.replace("Kids","Child")


		images_dict = {}
		for key, img in zip(["Main Image URL", "Additional Image URL (+)", "Additional Image URL 1 (+)", "Additional Image URL 2 (+)", "Additional Image URL 3 (+)"], images):
			images_dict[key] = img

		for key, p in zip(["Key Features (+)", "Key Features 1 (+)", "Key Features 2 (+)"], bullet_point):
			images_dict[key] = p

		for i,s in enumerate(stock):
			m_dt = {
				"Style Code": s["SKU"],
				"Product Name": new_title,
				"Product ID": s["Upc"],
				"Selling Price": price,
				"Total Count": s["Quantity"],
				"Brand": vendor,
				"Shipping Weight (lbs)": 3,

				"Shoe Size": s['size'],
				'Age Group (+)': age_group,
				'Shoe Category': f"{title_gender} Shoes",
				'Color Category (+)': color,

				'Total Count': s['Quantity'],
				"Variant Group ID": s['code'],
				"Gender":gender,
				"Site Description":description,
			}
			if i==0:
				m_dt["Is Primary Variant"]="Yes"
			m_dt.update(images_dict)
			ls = get_walmart_row(m_dt)
			# len(ls)
			sheet.append(ls)
		# break

	# if not sheet:
	# 	workbook.save('Walmart Product Data.xlsx')
	# 	workbook.close()
	# return sheet

def get_ebay_product_data(ls,vendor,workbook):
	# Replace with the path to your existing Excel file
	# file_path = 'TemplateEbayProduct.xlsx'
	# workbook =get_workbook("eBay")
	# Replace with the path to your existing Excel file
	# file_path = 'Templates/EbayProductTemplate.xlsx'
	# workbook = openpyxl.load_workbook(file_path)
	# sheet = workbook.active
	qty_sheet=workbook["eBay.Qty"]
	get_ebay_quantity_data(ls,vendor,qty_sheet)
	sheet=workbook["eBay"]
	for dt in ls:
		# Extracting data from dict
		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle = get_variables(dt,vendor)
		m_dt = {
			"Category name": category,
			"Title": new_title,
			"Relationship details": f"US Shoe Size={';'.join(map(str, sizes))}",
			"Start price": price,
			"Start price": '',
			"Item photo URL": "|".join(images),
			"C:Department": "Men",
			"C:Color": color,
			"Description":description
		}
		ls = get_ebay_row(m_dt, False)
		sheet.append(ls)
		for s in stock:
			m_dt = {
				"Custom label (SKU)": s["SKU"],
				"Title": new_title,
				"Relationship details": f"US Shoe Size={s['size']}",
				"P:UPC": s["Upc"],
				"Start price": price,
				"Relationship": "Variation",
				"Quantity": s["Quantity"]
			}
			ls = get_ebay_row(m_dt)
			sheet.append(ls)

		# workbook.save('Ebay Product Data.xlsx')
		# workbook.close()


def get_amazon_product_data(ls,vendor,workbook):
	# Replace with the path to your existing Excel file
	# file_path = 'TemplateEbayProduct.xlsx'
	# Replace with the path to your existing Excel file
	# file_path = 'Templates/AmazonProductTemplate.xlsx'
	# workbook = openpyxl.load_workbook(file_path)
	qty_sheet=workbook["Amazon.Qty"]
	get_amazon_quantity_data(ls,vendor,qty_sheet)
	sheet = workbook["Amazon"]
	for dt in ls:
		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle= get_variables(dt,vendor)


		images_dict = {}
		for key, img in zip(["Main Image URL","Other Image URL1","Other Image URL2","Other Image URL3","Other Image URL4","Other Image URL5","Other Image URL6","Other Image URL7","Other Image URL8"], images):
			images_dict[key] = img

		bullet_point_dict={}
		for key, b in zip(["Bullet Point1","Bullet Point2","Bullet Point3","Bullet Point4","Bullet Point5"], bullet_point):
			bullet_point_dict[key] = b

		if len(features)<2:
			features=["casual","cushioning"]+features
		feature_point_dict={}
		for key, f in zip(["Special Features1","Special Features2","Special Features3","Special Features4","Special Features5"],features):
			feature_point_dict[key] = f

		m_dt={
			"Product Type":type_,
			"Seller SKU":style_code,
			"Brand":vendor,
			"Product Name":new_title,
			"Product ID":"",
			"Product ID Type":"",
			"Product Description":description,
			"Shoe Size Gender":title_gender.replace('\'','').replace('\"',''),
			"Shoe Size Age Group":age_group,
			"Age Range Description":age_group,
			"Shoe Size":"",
			"Your Price":"",
			"Quantity":"",
			"Target Gender":"",
			"Age Range":"",
			"Description":"",
			"Parentage":"Parent",
			"Parent SKU":"",
			"Relationship Type":"",
			# "Color":main_color,
			"Color":color,
			"Color Map":color
			}
		m_dt.update(images_dict)
		m_dt.update(bullet_point_dict)
		m_dt.update(feature_point_dict)
		ls = get_amazon_product_row(m_dt)
		sheet.append(ls)
		for s in dt["stock"]:
			m_dt = {
			"Product Type":type_,
			"Seller SKU":s["SKU"],
			"Brand":vendor,
			"Product Name":new_title,
			"Product ID":s["Upc"],
			"Product ID Type":"UPC",
			"Shoe Size Age Group":gender,
			"Shoe Size Gender":title_gender.replace('\'','').replace('\"',''),
			"Shoe Size":s["size"],
			"Your Price":price,
			"Quantity":s["Quantity"],
			"Target Gender":gender,
			"Age Range":age_group,
			"Description":"",
			"Parentage":"Child",
			"Parent SKU":style_code,
			"Relationship Type":"Variation",
			"List Price":price,
			# "Color":main_color,
			"Color":color,
			"Color Map":color
			}
			ls = get_amazon_product_row(m_dt)
			sheet.append(ls)

	# workbook.save('Amazon Product Data.xlsx')
	# workbook.close()


def get_amazon_quantity_data(ls,vendor,sheet):
	# Replace with the path to your existing Excel file
	# file_path = 'TemplateEbayProduct.xlsx'
	# Replace with the path to your existing Excel file
	# file_path = 'Templates/AmazonQuantityTemplate.xlsx'
	# workbook = openpyxl.load_workbook(file_path)
	# sheet = workbook.active

	for dt in ls:

		url, title_gender, gender, age_group, title, style_code, color, sizes, price, cost, description, features, bullet_point, widths, images, category, weight, new_title,type_,main_color,stock,handle= get_variables(dt,vendor)
		images_dict = {}
		for key, img in zip(["Main Image URL", "Additional Image URL (+)", "Additional Image URL 1 (+)", "Additional Image URL 2 (+)", "Additional Image URL 3 (+)"], images):
			images_dict[key] = img

		for s in dt["stock"]:
			m_dt = {
				"sku": f"{s['SKU']}",
				"price": price,
				"minimum-seller-allowed-price": "",
				"maximum-seller-allowed-price": "",
				"quantity": f"{s['Quantity']}",
				"leadtime-to-ship": "3",
				"fulfillment-channel": ""

			}
			# m_dt.update(images_dict)
			ls = get_amazon_quantity_row(m_dt)
			# print(ls,len(ls))
			sheet.append(ls)
		# break
	# workbook.save('Amazon Quantity Data.xlsx')
	# workbook.close()
